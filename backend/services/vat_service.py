"""
VAT Calculation Service
Handles VAT calculations for transactions based on category settings
"""

import sys
import os
from typing import Dict, List, Optional, Tuple
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from sqlalchemy.orm import Session
from models import Transaction, CustomCategory, SessionVATConfig, SessionLocal

# Import VAT defaults from categories_service to keep them in sync
try:
    from services.categories_service import BUILT_IN_CATEGORY_VAT_DEFAULTS
except ImportError:
    # Fallback if import fails
    BUILT_IN_CATEGORY_VAT_DEFAULTS = {
        "Fuel": {"applicable": True, "rate": 15.0},
        "Bank Fees": {"applicable": False, "rate": 0.0},
        "Rent": {"applicable": False, "rate": 0.0},
        "Salary": {"applicable": False, "rate": 0.0},
        "Groceries": {"applicable": True, "rate": 15.0},
        "Utilities": {"applicable": True, "rate": 15.0},
        "Transport": {"applicable": True, "rate": 15.0},
        "Healthcare": {"applicable": False, "rate": 0.0},
        "Insurance": {"applicable": False, "rate": 0.0},
        "Entertainment": {"applicable": True, "rate": 15.0},
        "Clothing": {"applicable": True, "rate": 15.0},
        "Dining": {"applicable": True, "rate": 15.0},
        "Travel": {"applicable": True, "rate": 15.0},
        "Education": {"applicable": False, "rate": 0.0},
        "Other": {"applicable": False, "rate": 0.0},
    }

# Income/Sales categories (VAT Output) - vs expenses (VAT Input)
INCOME_CATEGORIES = {"Salary", "Income", "Sales"}


class VATService:
    """Service for managing VAT calculations and configuration"""
    
    def __init__(self):
        pass
    
    def _get_db(self) -> Session:
        """Get a database session"""
        return SessionLocal()
    
    def _is_vat_output(self, category: str) -> bool:
        """Determine if a transaction is VAT Output (income/sales) or Input (expenses)
        
        Uses hybrid approach:
        1. Check hardcoded built-in income categories first (fast)
        2. Check custom categories in database (flexible)
        
        Args:
            category: Transaction category name
            
        Returns:
            True if VAT Output (income), False if VAT Input (expense)
        """
        # Fast path: Check built-in income categories
        if category in INCOME_CATEGORIES:
            return True
        
        # Slow path: Check custom categories in database
        db = self._get_db()
        try:
            custom_cat = db.query(CustomCategory).filter(
                CustomCategory.name == category
            ).first()
            
            if custom_cat:
                return custom_cat.is_income == 1
            
            # Not found - default to VAT Input (expense)
            return False
        finally:
            db.close()
    
    def _split_transactions_by_type(self, transactions: List[Transaction]) -> Tuple[List[Transaction], List[Transaction]]:
        """Split transactions into VAT Input (expenses) and VAT Output (income/sales)
        
        Args:
            transactions: List of transactions
            
        Returns:
            Tuple of (vat_input_transactions, vat_output_transactions)
        """
        vat_input = []
        vat_output = []
        
        for txn in transactions:
            if self._is_vat_output(txn.category):
                vat_output.append(txn)
            else:
                vat_input.append(txn)
        
        return vat_input, vat_output
    
    def get_session_vat_config(self, session_id: str) -> Optional[SessionVATConfig]:
        """Get VAT configuration for a session"""
        db = self._get_db()
        try:
            config = db.query(SessionVATConfig).filter(
                SessionVATConfig.session_id == session_id
            ).first()
            return config
        finally:
            db.close()
    
    def is_vat_enabled(self, session_id: str) -> bool:
        """Check if VAT is enabled for a session"""
        config = self.get_session_vat_config(session_id)
        return config.vat_enabled == 1 if config else False
    
    def enable_vat(self, session_id: str, default_rate: float = 15.0) -> Tuple[bool, str]:
        """Enable VAT calculation for a session"""
        db = self._get_db()
        try:
            config = db.query(SessionVATConfig).filter(
                SessionVATConfig.session_id == session_id
            ).first()
            
            if config:
                config.vat_enabled = 1
                config.default_vat_rate = default_rate
                config.updated_at = datetime.utcnow()
            else:
                config = SessionVATConfig(
                    session_id=session_id,
                    vat_enabled=1,
                    default_vat_rate=default_rate
                )
                db.add(config)
            
            db.commit()
            return True, "VAT calculation enabled"
        except Exception as e:
            db.rollback()
            return False, f"Failed to enable VAT: {str(e)}"
        finally:
            db.close()
    
    def disable_vat(self, session_id: str) -> Tuple[bool, str]:
        """Disable VAT calculation for a session"""
        db = self._get_db()
        try:
            config = db.query(SessionVATConfig).filter(
                SessionVATConfig.session_id == session_id
            ).first()
            
            if config:
                config.vat_enabled = 0
                config.updated_at = datetime.utcnow()
                db.commit()
            
            return True, "VAT calculation disabled"
        except Exception as e:
            db.rollback()
            return False, f"Failed to disable VAT: {str(e)}"
        finally:
            db.close()
    
    def get_category_vat_settings(self, category_name: str) -> Dict[str, any]:
        """Get VAT settings for a category (built-in or custom)"""
        # Check if it's a custom category first
        db = self._get_db()
        try:
            custom_cat = db.query(CustomCategory).filter(
                CustomCategory.name == category_name
            ).first()
            
            if custom_cat:
                return {
                    "applicable": custom_cat.vat_applicable == 1,
                    "rate": custom_cat.vat_rate
                }
            
            # Check built-in categories
            if category_name in BUILT_IN_CATEGORY_VAT_DEFAULTS:
                return BUILT_IN_CATEGORY_VAT_DEFAULTS[category_name]
            
            # Default for unknown categories
            return {"applicable": False, "rate": 0.0}
        finally:
            db.close()
    
    def update_category_vat_settings(
        self,
        category_name: str,
        vat_applicable: bool,
        vat_rate: float,
        is_income: Optional[bool] = None
    ) -> Tuple[bool, str]:
        """
        Update VAT settings for any category (built-in or custom)
        For built-in categories, create a custom category entry to override defaults
        
        Args:
            category_name: Name of the category
            vat_applicable: Whether VAT applies to this category
            vat_rate: VAT rate percentage
            is_income: Optional - True for Income/Sales (VAT Output), False for Expense (VAT Input)
        """
        db = self._get_db()
        try:
            # Check if custom category already exists
            custom_cat = db.query(CustomCategory).filter(
                CustomCategory.name == category_name
            ).first()
            
            if custom_cat:
                # Update existing custom category
                custom_cat.vat_applicable = 1 if vat_applicable else 0
                custom_cat.vat_rate = vat_rate
                if is_income is not None:
                    custom_cat.is_income = 1 if is_income else 0
            else:
                # Create new custom category entry (for built-in category overrides)
                custom_cat = CustomCategory(
                    name=category_name,
                    vat_applicable=1 if vat_applicable else 0,
                    vat_rate=vat_rate,
                    is_income=1 if (is_income is not None and is_income) else 0
                )
                db.add(custom_cat)
            
            db.commit()
            return True, f"VAT settings updated for {category_name}"
        except Exception as e:
            db.rollback()
            return False, f"Failed to update VAT settings: {str(e)}"
        finally:
            db.close()
    
    def calculate_vat(
        self,
        amount: float,
        vat_rate: float,
        amount_includes_vat: bool = True
    ) -> Dict[str, float]:
        """
        Calculate VAT amounts
        
        Args:
            amount: The transaction amount
            vat_rate: VAT rate percentage (e.g., 15 for 15%)
            amount_includes_vat: If True, amount is VAT inclusive; if False, VAT exclusive
        
        Returns:
            Dict with vat_amount, amount_excl_vat, amount_incl_vat
        """
        if amount_includes_vat:
            # Amount is VAT inclusive (most common for purchases)
            # Formula: VAT = Amount × (rate / (100 + rate))
            vat_amount = amount * (vat_rate / (100 + vat_rate))
            amount_excl_vat = amount - vat_amount
            amount_incl_vat = amount
        else:
            # Amount is VAT exclusive (less common)
            # Formula: VAT = Amount × (rate / 100)
            vat_amount = amount * (vat_rate / 100)
            amount_excl_vat = amount
            amount_incl_vat = amount + vat_amount
        
        return {
            "vat_amount": round(vat_amount, 2),
            "amount_excl_vat": round(amount_excl_vat, 2),
            "amount_incl_vat": round(amount_incl_vat, 2)
        }
    
    def apply_vat_to_transaction(
        self,
        transaction_id: int,
        session_id: str,
        force: bool = False
    ) -> Tuple[bool, str]:
        """
        Calculate and apply VAT to a transaction based on its category
        
        Args:
            transaction_id: ID of the transaction
            session_id: Session ID
            force: If True, apply even if VAT is disabled
        """
        # Check if VAT is enabled for this session
        if not force and not self.is_vat_enabled(session_id):
            return False, "VAT calculation is not enabled for this session"
        
        db = self._get_db()
        try:
            transaction = db.query(Transaction).filter(
                Transaction.id == transaction_id,
                Transaction.session_id == session_id
            ).first()
            
            if not transaction:
                return False, "Transaction not found"
            
            # Get VAT settings for the transaction's category
            vat_settings = self.get_category_vat_settings(transaction.category)
            
            if not vat_settings["applicable"]:
                # Clear VAT fields if not applicable
                transaction.vat_amount = None
                transaction.amount_excl_vat = None
                transaction.amount_incl_vat = None
            else:
                # Calculate VAT (assuming amount is VAT inclusive)
                vat_calc = self.calculate_vat(
                    transaction.amount,
                    vat_settings["rate"],
                    amount_includes_vat=True
                )
                
                transaction.vat_amount = vat_calc["vat_amount"]
                transaction.amount_excl_vat = vat_calc["amount_excl_vat"]
                transaction.amount_incl_vat = vat_calc["amount_incl_vat"]
            
            db.commit()
            return True, "VAT calculated successfully"
        except Exception as e:
            db.rollback()
            return False, f"Failed to calculate VAT: {str(e)}"
        finally:
            db.close()
    
    def recalculate_all_transactions(self, session_id: str) -> Tuple[bool, str, Dict]:
        """Recalculate VAT for all transactions in a session"""
        if not self.is_vat_enabled(session_id):
            return False, "VAT calculation is not enabled for this session", {}
        
        db = self._get_db()
        try:
            transactions = db.query(Transaction).filter(
                Transaction.session_id == session_id
            ).all()
            
            updated_count = 0
            skipped_count = 0
            
            for transaction in transactions:
                vat_settings = self.get_category_vat_settings(transaction.category)
                
                if not vat_settings["applicable"]:
                    # Clear VAT fields
                    transaction.vat_amount = None
                    transaction.amount_excl_vat = None
                    transaction.amount_incl_vat = None
                    skipped_count += 1
                else:
                    # Calculate VAT
                    vat_calc = self.calculate_vat(
                        transaction.amount,
                        vat_settings["rate"],
                        amount_includes_vat=True
                    )
                    
                    transaction.vat_amount = vat_calc["vat_amount"]
                    transaction.amount_excl_vat = vat_calc["amount_excl_vat"]
                    transaction.amount_incl_vat = vat_calc["amount_incl_vat"]
                    updated_count += 1
            
            db.commit()
            
            return True, "VAT recalculated for all transactions", {
                "total": len(transactions),
                "updated": updated_count,
                "skipped": skipped_count
            }
        except Exception as e:
            db.rollback()
            return False, f"Failed to recalculate VAT: {str(e)}", {}
        finally:
            db.close()
    
    def get_vat_summary(self, session_id: str, start_date: Optional[date] = None, end_date: Optional[date] = None) -> Dict:
        """
        Get VAT summary for a session
        
        Returns:
            Dict with total_vat, total_excl_vat, total_incl_vat, transactions_count
        """
        db = self._get_db()
        try:
            query = db.query(Transaction).filter(
                Transaction.session_id == session_id,
                Transaction.vat_amount.isnot(None)
            )
            
            if start_date:
                query = query.filter(Transaction.date >= start_date)
            if end_date:
                query = query.filter(Transaction.date <= end_date)
            
            transactions = query.all()
            
            total_vat = sum(t.vat_amount or 0 for t in transactions)
            total_excl_vat = sum(t.amount_excl_vat or 0 for t in transactions)
            total_incl_vat = sum(t.amount_incl_vat or 0 for t in transactions)
            
            # Group by category
            by_category = {}
            for t in transactions:
                if t.category not in by_category:
                    by_category[t.category] = {
                        "count": 0,
                        "total_vat": 0.0,
                        "total_excl_vat": 0.0,
                        "total_incl_vat": 0.0
                    }
                
                by_category[t.category]["count"] += 1
                by_category[t.category]["total_vat"] += t.vat_amount or 0
                by_category[t.category]["total_excl_vat"] += t.amount_excl_vat or 0
                by_category[t.category]["total_incl_vat"] += t.amount_incl_vat or 0
            
            return {
                "total_vat": round(total_vat, 2),
                "total_excl_vat": round(total_excl_vat, 2),
                "total_incl_vat": round(total_incl_vat, 2),
                "transactions_count": len(transactions),
                "by_category": by_category,
                "period": {
                    "start_date": start_date.isoformat() if start_date else None,
                    "end_date": end_date.isoformat() if end_date else None
                }
            }
        finally:
            db.close()
    
    def export_vat_report(
        self,
        session_id: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        format: str = "excel",
        client_id: Optional[int] = None,
        export_type: str = "both"
    ):
        """
        Export VAT report in Excel or CSV format
        
        Args:
            session_id: Optional Session ID
            start_date: Optional start date filter
            end_date: Optional end date filter
            format: "excel" or "csv"
            client_id: Optional Client ID (alternative to session_id)
            export_type: "both", "input_only", or "output_only"
        
        Returns:
            BytesIO object with the report
        """
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import csv
        from models import TransactionMerchant
        
        db = self._get_db()
        try:
            # Get transactions with VAT
            query = db.query(Transaction).filter(
                Transaction.vat_amount.isnot(None)
            )
            
            # Filter by session_id or client_id
            if session_id:
                query = query.filter(Transaction.session_id == session_id)
            elif client_id:
                query = query.filter(Transaction.client_id == client_id)
            
            if start_date:
                query = query.filter(Transaction.date >= start_date)
            if end_date:
                query = query.filter(Transaction.date <= end_date)
            
            transactions = query.order_by(Transaction.date).all()
            
            # Fetch merchants for all transactions
            for txn in transactions:
                merchant_record = db.query(TransactionMerchant).filter(
                    TransactionMerchant.transaction_id == txn.id
                ).first()
                txn.merchant = merchant_record.merchant if merchant_record else None
            
            if format == "csv":
                return self._export_vat_csv(transactions, start_date, end_date, export_type)
            else:
                return self._export_vat_excel(transactions, start_date, end_date, export_type)
        finally:
            db.close()
    
    def _export_vat_csv(self, transactions: List[Transaction], start_date, end_date, export_type: str = "both"):
        """Export VAT report as CSV with separate sections for Input and Output"""
        from io import StringIO, BytesIO
        import csv
        
        # Split transactions into input and output
        vat_input, vat_output = self._split_transactions_by_type(transactions)
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        type_label = {
            'both': 'VAT REPORT',
            'input_only': 'VAT INPUT REPORT',
            'output_only': 'VAT OUTPUT REPORT'
        }
        writer.writerow([type_label.get(export_type, 'VAT REPORT')])
        if start_date and end_date:
            writer.writerow([f"Period: {start_date} to {end_date}"])
        writer.writerow([])
        
        # Track totals
        input_incl = input_vat = input_excl = 0.0
        output_incl = output_vat = output_excl = 0.0
        
        # VAT INPUT (Expenses) Section
        if export_type in ['both', 'input_only']:
            writer.writerow(["VAT INPUT (Expenses) - VAT Claimable"])
            writer.writerow([
                "Date",
                "Description",
                "Category",
                "Merchant",
                "Amount (Incl VAT)",
                "VAT Amount",
                "Amount (Excl VAT)"
            ])
            
            for txn in vat_input:
                writer.writerow([
                    txn.date.isoformat(),
                    txn.description,
                    txn.category,
                    txn.merchant or "",
                    f"{txn.amount_incl_vat:.2f}",
                    f"{txn.vat_amount:.2f}",
                    f"{txn.amount_excl_vat:.2f}"
                ])
                input_incl += txn.amount_incl_vat or 0
                input_vat += txn.vat_amount or 0
                input_excl += txn.amount_excl_vat or 0
            
            # VAT Input Subtotals
            writer.writerow([])
            writer.writerow([
                "INPUT SUBTOTALS",
                "",
                "",
                "",
                f"{input_incl:.2f}",
                f"{input_vat:.2f}",
                f"{input_excl:.2f}"
            ])
            writer.writerow([])
        
        # VAT OUTPUT (Income/Sales) Section
        if export_type in ['both', 'output_only']:
            if export_type == 'both':
                writer.writerow([])
            writer.writerow(["VAT OUTPUT (Sales/Income) - VAT Payable"])
            writer.writerow([
                "Date",
                "Description",
                "Category",
                "Merchant",
                "Amount (Incl VAT)",
                "VAT Amount",
                "Amount (Excl VAT)"
            ])
            
            for txn in vat_output:
                writer.writerow([
                    txn.date.isoformat(),
                    txn.description,
                    txn.category,
                    txn.merchant or "",
                    f"{txn.amount_incl_vat:.2f}",
                    f"{txn.vat_amount:.2f}",
                    f"{txn.amount_excl_vat:.2f}"
                ])
                output_incl += txn.amount_incl_vat or 0
                output_vat += txn.vat_amount or 0
                output_excl += txn.amount_excl_vat or 0
            
            # VAT Output Subtotals
            writer.writerow([])
            writer.writerow([
                "OUTPUT SUBTOTALS",
                "",
                "",
                "",
                f"{output_incl:.2f}",
                f"{output_vat:.2f}",
                f"{output_excl:.2f}"
            ])
            writer.writerow([])
        
        # Overall Totals and Summary
        writer.writerow([])
        writer.writerow(["VAT SUMMARY"])
        
        if export_type == 'both':
            writer.writerow(["Total Transactions", len(transactions)])
            writer.writerow(["  - Input (Expenses)", len(vat_input)])
            writer.writerow(["  - Output (Sales/Income)", len(vat_output)])
        elif export_type == 'input_only':
            writer.writerow(["Total Input Transactions", len(vat_input)])
        else:  # output_only
            writer.writerow(["Total Output Transactions", len(vat_output)])
        
        writer.writerow([])
        
        # Net VAT only for 'both'
        if export_type == 'both':
            writer.writerow(["NET VAT POSITION"])
            writer.writerow(["VAT Claimable (Input)", f"R {abs(input_vat):.2f}"])
            writer.writerow(["VAT Payable (Output)", f"R {abs(output_vat):.2f}"])
            net_vat = abs(input_vat) - abs(output_vat)
            writer.writerow(["Net VAT (To Claim/Pay)", f"R {net_vat:.2f}"])
            writer.writerow([])
            writer.writerow(["TOTAL AMOUNTS"])
            writer.writerow(["Total Input (Incl VAT)", f"R {input_incl:.2f}"])
            writer.writerow(["Total Output (Incl VAT)", f"R {output_incl:.2f}"])
            writer.writerow(["Grand Total (Incl VAT)", f"R {input_incl + output_incl:.2f}"])
        elif export_type == 'input_only':
            writer.writerow(["TOTAL VAT CLAIMABLE (INPUT)"])
            writer.writerow(["VAT Amount", f"R {abs(input_vat):.2f}"])
            writer.writerow(["Total (Incl VAT)", f"R {input_incl:.2f}"])
        else:  # output_only
            writer.writerow(["TOTAL VAT PAYABLE (OUTPUT)"])
            writer.writerow(["VAT Amount", f"R {abs(output_vat):.2f}"])
            writer.writerow(["Total (Incl VAT)", f"R {output_incl:.2f}"])
        
        # Convert to bytes
        bytes_output = BytesIO(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)
        return bytes_output
    
    def _export_vat_excel(self, transactions: List[Transaction], start_date, end_date, export_type: str = "both"):
        """Export VAT report as Excel with separate sections for Input and Output"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        # Split transactions into input and output
        vat_input, vat_output = self._split_transactions_by_type(transactions)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VAT Report"
        
        # Styling
        HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        SECTION_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
        SUBTOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        SUBTOTAL_FONT = Font(bold=True, size=11)
        TOTAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        TOTAL_FONT = Font(bold=True, size=12)
        BORDER = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        current_row = 1
        
        # Title
        type_label = {
            'both': 'VAT REPORT',
            'input_only': 'VAT INPUT REPORT',
            'output_only': 'VAT OUTPUT REPORT'
        }
        ws.cell(row=current_row, column=1).value = type_label.get(export_type, 'VAT REPORT')
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        # Period
        if start_date and end_date:
            ws.cell(row=current_row, column=1).value = f"Period: {start_date} to {end_date}"
            ws.cell(row=current_row, column=1).font = Font(italic=True)
            current_row += 2
        else:
            current_row += 1
        
        # Helper function to write a transaction section
        def write_section(section_title: str, transactions_list: List[Transaction], row_num: int):
            # Section header
            ws.cell(row=row_num, column=1).value = section_title
            ws.cell(row=row_num, column=1).font = SECTION_FONT
            ws.cell(row=row_num, column=1).fill = SECTION_FILL
            ws.merge_cells(f'A{row_num}:G{row_num}')
            row_num += 1
            
            # Column headers
            headers = [
                "Date",
                "Description",
                "Category",
                "Merchant",
                "Amount (Incl VAT)",
                "VAT Amount",
                "Amount (Excl VAT)"
            ]
            
            header_row = row_num
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
                cell.border = BORDER
            
            # Data rows
            data_start_row = header_row + 1
            section_incl = 0.0
            section_vat = 0.0
            section_excl = 0.0
            
            for idx, txn in enumerate(transactions_list, data_start_row):
                ws.cell(row=idx, column=1).value = txn.date
                ws.cell(row=idx, column=1).number_format = 'yyyy-mm-dd'
                ws.cell(row=idx, column=2).value = txn.description
                ws.cell(row=idx, column=3).value = txn.category
                ws.cell(row=idx, column=4).value = txn.merchant or ""
                ws.cell(row=idx, column=5).value = txn.amount_incl_vat
                ws.cell(row=idx, column=5).number_format = 'R #,##0.00'
                ws.cell(row=idx, column=6).value = txn.vat_amount
                ws.cell(row=idx, column=6).number_format = 'R #,##0.00'
                ws.cell(row=idx, column=7).value = txn.amount_excl_vat
                ws.cell(row=idx, column=7).number_format = 'R #,##0.00'
                
                for col in range(1, 8):
                    ws.cell(row=idx, column=col).border = BORDER
                
                section_incl += txn.amount_incl_vat or 0
                section_vat += txn.vat_amount or 0
                section_excl += txn.amount_excl_vat or 0
            
            # Subtotals row
            subtotal_row = data_start_row + len(transactions_list)
            ws.cell(row=subtotal_row, column=1).value = "SUBTOTALS"
            ws.cell(row=subtotal_row, column=1).font = SUBTOTAL_FONT
            ws.cell(row=subtotal_row, column=1).fill = SUBTOTAL_FILL
            
            for col in range(2, 5):
                ws.cell(row=subtotal_row, column=col).fill = SUBTOTAL_FILL
            
            ws.cell(row=subtotal_row, column=5).value = section_incl
            ws.cell(row=subtotal_row, column=5).number_format = 'R #,##0.00'
            ws.cell(row=subtotal_row, column=5).font = SUBTOTAL_FONT
            ws.cell(row=subtotal_row, column=5).fill = SUBTOTAL_FILL
            
            ws.cell(row=subtotal_row, column=6).value = section_vat
            ws.cell(row=subtotal_row, column=6).number_format = 'R #,##0.00'
            ws.cell(row=subtotal_row, column=6).font = SUBTOTAL_FONT
            ws.cell(row=subtotal_row, column=6).fill = SUBTOTAL_FILL
            
            ws.cell(row=subtotal_row, column=7).value = section_excl
            ws.cell(row=subtotal_row, column=7).number_format = 'R #,##0.00'
            ws.cell(row=subtotal_row, column=7).font = SUBTOTAL_FONT
            ws.cell(row=subtotal_row, column=7).fill = SUBTOTAL_FILL
            
            for col in range(1, 8):
                ws.cell(row=subtotal_row, column=col).border = BORDER
            
            return subtotal_row + 1, section_incl, section_vat, section_excl
        
        # Track totals
        input_incl = input_vat = input_excl = 0.0
        output_incl = output_vat = output_excl = 0.0
        
        # Write VAT INPUT section
        if export_type in ['both', 'input_only']:
            current_row, input_incl, input_vat, input_excl = write_section(
                "VAT INPUT (Expenses) - VAT Claimable",
                vat_input,
                current_row
            )
            current_row += 2
        
        # Write VAT OUTPUT section
        if export_type in ['both', 'output_only']:
            current_row, output_incl, output_vat, output_excl = write_section(
                "VAT OUTPUT (Sales/Income) - VAT Payable",
                vat_output,
                current_row
            )
            current_row += 3
        
        # Summary section
        ws.cell(row=current_row, column=1).value = "VAT SUMMARY"
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1
        
        if export_type == 'both':
            ws.cell(row=current_row, column=1).value = "Total Transactions:"
            ws.cell(row=current_row, column=2).value = len(transactions)
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "  - Input (Expenses):"
            ws.cell(row=current_row, column=2).value = len(vat_input)
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "  - Output (Sales/Income):"
            ws.cell(row=current_row, column=2).value = len(vat_output)
            current_row += 2
            
            # NET VAT POSITION section (only for 'both')
            ws.cell(row=current_row, column=1).value = "NET VAT POSITION"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "VAT Claimable (Input):"
            ws.cell(row=current_row, column=2).value = abs(input_vat)
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "VAT Payable (Output):"
            ws.cell(row=current_row, column=2).value = abs(output_vat)
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            current_row += 1
            
            net_vat = abs(input_vat) - abs(output_vat)
            ws.cell(row=current_row, column=1).value = "Net VAT (To Claim/Pay):"
            ws.cell(row=current_row, column=2).value = net_vat
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = TOTAL_FONT
            ws.cell(row=current_row, column=2).fill = TOTAL_FILL
            current_row += 2
            
            # TOTAL AMOUNTS section
            ws.cell(row=current_row, column=1).value = "TOTAL AMOUNTS"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "Total Input (Incl VAT):"
            ws.cell(row=current_row, column=2).value = input_incl
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "Total Output (Incl VAT):"
            ws.cell(row=current_row, column=2).value = output_incl
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "Grand Total (Incl VAT):"
            ws.cell(row=current_row, column=2).value = input_incl + output_incl
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            
        elif export_type == 'input_only':
            ws.cell(row=current_row, column=1).value = "Total Input Transactions:"
            ws.cell(row=current_row, column=2).value = len(vat_input)
            current_row += 2
            
            ws.cell(row=current_row, column=1).value = "TOTAL VAT CLAIMABLE (INPUT)"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "VAT Amount:"
            ws.cell(row=current_row, column=2).value = abs(input_vat)
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = TOTAL_FONT
            ws.cell(row=current_row, column=2).fill = TOTAL_FILL
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "Total (Incl VAT):"
            ws.cell(row=current_row, column=2).value = input_incl
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            
        else:  # output_only
            ws.cell(row=current_row, column=1).value = "Total Output Transactions:"
            ws.cell(row=current_row, column=2).value = len(vat_output)
            current_row += 2
            
            ws.cell(row=current_row, column=1).value = "TOTAL VAT PAYABLE (OUTPUT)"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "VAT Amount:"
            ws.cell(row=current_row, column=2).value = abs(output_vat)
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = TOTAL_FONT
            ws.cell(row=current_row, column=2).fill = TOTAL_FILL
            current_row += 1
            
            ws.cell(row=current_row, column=1).value = "Total (Incl VAT):"
            ws.cell(row=current_row, column=2).value = output_incl
            ws.cell(row=current_row, column=2).number_format = 'R #,##0.00'
            ws.cell(row=current_row, column=2).font = Font(bold=True)
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
