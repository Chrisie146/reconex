"""
Summary Service
Calculates monthly summaries and aggregations for reporting
"""

from typing import Dict, List, Any, Optional
from datetime import date
from sqlalchemy.orm import Session, joinedload
from models import Account, Transaction


def calculate_monthly_summary(session_id: str = None, db: Session = None, client_id: int = None) -> Dict[str, Any]:
    """
    Calculate monthly summary for all transactions in a session or by client
    
    Returns:
    {
        "months": [
            {
                "month": "2024-01",
                "total_income": 5000.00,
                "total_expenses": 3500.00,
                "net_balance": 1500.00,
                "categories": {
                    "Rent": 1500.00,
                    "Utilities": 250.00,
                    ...
                }
            },
            ...
        ],
        "overall": {
            "total_income": 15000.00,
            "total_expenses": 10500.00,
            "net_balance": 4500.00
        }
    }
    """
    
    # Fetch transactions - filter by session_id or client_id
    query = db.query(Transaction).options(joinedload(Transaction.account))
    if session_id:
        query = query.filter(Transaction.session_id == session_id)
    elif client_id:
        query = query.filter(Transaction.client_id == client_id)
    else:
        raise ValueError("Either session_id or client_id must be provided")

    transactions = query.order_by(Transaction.date).all()

    if not transactions:
        return {
            "months": [],
            "overall": {
                "total_income": 0.0,
                "total_expenses": 0.0,
                "net_balance": 0.0,
            },
        }

    # Group by month
    monthly_data: Dict[str, Dict[str, Any]] = {}

    for txn in transactions:
        month_key = txn.date.strftime("%Y-%m")

        if month_key not in monthly_data:
            monthly_data[month_key] = {
                "month": month_key,
                "total_income": 0.0,
                "total_expenses": 0.0,
                "net_balance": 0.0,
                "categories": {},
                "accounts": {},  # Phase 2
            }

        month = monthly_data[month_key]

        # Income vs expense classification: prefer account_type when set
        acct = txn.account
        if acct and txn.account_id:
            is_income = acct.account_type in ("revenue",)
        else:
            is_income = txn.amount >= 0

        if is_income:
            month["total_income"] += abs(txn.amount)
        else:
            month["total_expenses"] += abs(txn.amount)

        # Category grouping (legacy, always populated)
        cat_key = txn.category or "Uncategorized"
        if cat_key not in month["categories"]:
            month["categories"][cat_key] = 0.0
        month["categories"][cat_key] += abs(txn.amount)

        # Account grouping (Phase 2 — only for transactions with account_id)
        if acct and txn.account_id:
            acct_key = f"{acct.code} · {acct.name}"
            if acct_key not in month["accounts"]:
                month["accounts"][acct_key] = {
                    "account_id": acct.id,
                    "code": acct.code,
                    "name": acct.name,
                    "account_type": acct.account_type,
                    "total": 0.0,
                }
            month["accounts"][acct_key]["total"] += abs(txn.amount)

        month["net_balance"] = month["total_income"] - month["total_expenses"]

    sorted_months = sorted(monthly_data.items())
    months_list = [data for _, data in sorted_months]

    overall_income = sum(m["total_income"] for m in months_list)
    overall_expenses = sum(m["total_expenses"] for m in months_list)
    overall_balance = overall_income - overall_expenses

    return {
        "months": months_list,
        "overall": {
            "total_income": overall_income,
            "total_expenses": overall_expenses,
            "net_balance": overall_balance,
        },
    }


def get_category_summary(session_id: str = None, db: Session = None, client_id: int = None) -> Dict[str, float]:
    """Get total amounts by category across all transactions in a session or by client."""
    query = db.query(Transaction)
    if session_id:
        query = query.filter(Transaction.session_id == session_id)
    elif client_id:
        query = query.filter(Transaction.client_id == client_id)
    else:
        raise ValueError("Either session_id or client_id must be provided")

    transactions = query.all()

    categories: Dict[str, float] = {}

    for txn in transactions:
        key = txn.category or "Uncategorized"
        if key not in categories:
            categories[key] = 0.0
        categories[key] += abs(txn.amount)

    return dict(sorted(categories.items(), key=lambda x: x[1], reverse=True))


def get_account_summary(
    session_id: Optional[str] = None,
    db: Session = None,
    client_id: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """Get total amounts grouped by Chart-of-Accounts account.

    Phase 2: only includes transactions that have an account_id assigned.
    Transactions without an account_id are not included — use
    get_category_summary() for those.

    Returns:
        List of dicts ordered by account.code, each with:
        { account_id, code, name, account_type, normal_balance, total }
    """
    query = (
        db.query(Transaction)
        .options(joinedload(Transaction.account))
        .filter(Transaction.account_id.isnot(None))
    )
    if session_id:
        query = query.filter(Transaction.session_id == session_id)
    elif client_id:
        query = query.filter(Transaction.client_id == client_id)
    else:
        raise ValueError("Either session_id or client_id must be provided")

    transactions = query.all()

    accounts: Dict[int, Dict[str, Any]] = {}
    for txn in transactions:
        acct = txn.account
        if not acct:
            continue
        if acct.id not in accounts:
            accounts[acct.id] = {
                "account_id": acct.id,
                "code": acct.code,
                "name": acct.name,
                "account_type": acct.account_type,
                "normal_balance": acct.normal_balance,
                "total": 0.0,
            }
        accounts[acct.id]["total"] += abs(txn.amount)

    return sorted(accounts.values(), key=lambda x: x["code"] or "")


def get_transactions_by_category(
    category: str,
    db: Session,
    session_id: Optional[str] = None,
    client_id: Optional[int] = None
) -> List[Dict[str, Any]]:
    """
    Get all transactions for a specific category
    Accepts either session_id or client_id
    """
    
    query = db.query(Transaction).filter(Transaction.category == category)
    
    if session_id:
        query = query.filter(Transaction.session_id == session_id)
    elif client_id:
        query = query.filter(Transaction.client_id == client_id)
    else:
        raise ValueError("Either session_id or client_id must be provided")
    
    transactions = query.order_by(Transaction.date.desc()).all()
    
    return [
        {
            "id": txn.id,
            "date": txn.date.isoformat(),
            "description": txn.description,
            "amount": txn.amount,
            "category": txn.category,
            "invoice_id": txn.invoice_id
        }
        for txn in transactions
    ]


# Excel Export Implementation
import os
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from models import TransactionMerchant


class ExcelExporter:
    """Generate polished, accountant-ready Excel exports with branded styling."""

    # ── Colour palette (indigo brand) ──────────────────────────
    _INDIGO_DARK   = "3730A3"   # indigo-800  – header backgrounds
    _INDIGO_MID    = "4F46E5"   # indigo-600  – accent / titles
    _INDIGO_LIGHT  = "E0E7FF"   # indigo-100  – alternate row stripe
    _INDIGO_PALE   = "EEF2FF"   # indigo-50   – very light tint
    _NEUTRAL_100   = "F5F5F5"
    _NEUTRAL_200   = "E5E5E5"
    _NEUTRAL_700   = "404040"
    _GREEN_DARK    = "065F46"   # emerald-800
    _GREEN_BG      = "D1FAE5"   # emerald-100
    _RED_DARK      = "991B1B"   # red-800
    _RED_BG        = "FEE2E2"   # red-100

    # ── Reusable style objects ─────────────────────────────────
    HEADER_FILL  = PatternFill(start_color=_INDIGO_DARK, end_color=_INDIGO_DARK, fill_type="solid")
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    TOTAL_FILL   = PatternFill(start_color=_INDIGO_LIGHT, end_color=_INDIGO_LIGHT, fill_type="solid")
    TOTAL_FONT   = Font(name="Calibri", bold=True, size=10)
    STRIPE_FILL  = PatternFill(start_color=_INDIGO_PALE, end_color=_INDIGO_PALE, fill_type="solid")
    SUBTITLE_FILL = PatternFill(start_color=_NEUTRAL_200, end_color=_NEUTRAL_200, fill_type="solid")
    SUBTITLE_FONT = Font(name="Calibri", bold=True, size=10, color=_NEUTRAL_700)
    TITLE_FONT   = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    BODY_FONT    = Font(name="Calibri", size=10)
    BORDER = Border(
        left=Side(style="thin", color=_NEUTRAL_200),
        right=Side(style="thin", color=_NEUTRAL_200),
        top=Side(style="thin", color=_NEUTRAL_200),
        bottom=Side(style="thin", color=_NEUTRAL_200),
    )
    CURRENCY_FMT = 'R #,##0.00'
    AMOUNT_POSITIVE_FONT = Font(name="Calibri", size=10, color=_GREEN_DARK)
    AMOUNT_NEGATIVE_FONT = Font(name="Calibri", size=10, color=_RED_DARK)

    # ── Helpers ────────────────────────────────────────────────
    @staticmethod
    def _style_header_row(ws, row: int, num_cols: int):
        """Apply branded header style to a row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ExcelExporter.BORDER

    @staticmethod
    def _style_data_row(ws, row: int, num_cols: int, is_stripe: bool = False):
        """Apply body/stripe style to a data row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = ExcelExporter.BODY_FONT
            cell.border = ExcelExporter.BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if is_stripe:
                cell.fill = ExcelExporter.STRIPE_FILL

    @staticmethod
    def _style_total_row(ws, row: int, num_cols: int):
        """Apply totals bar style."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = ExcelExporter.TOTAL_FILL
            cell.font = ExcelExporter.TOTAL_FONT
            cell.border = ExcelExporter.BORDER

    @staticmethod
    def _style_amount_cell(cell, value):
        """Colour an amount cell green (positive) or red (negative)."""
        cell.number_format = ExcelExporter.CURRENCY_FMT
        if value is not None:
            if isinstance(value, (int, float)):
                cell.font = (
                    ExcelExporter.AMOUNT_POSITIVE_FONT if value >= 0
                    else ExcelExporter.AMOUNT_NEGATIVE_FONT
                )

    @staticmethod
    def _freeze_and_filter(ws, header_row: int, num_cols: int):
        """Freeze panes below the header and add auto-filter."""
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        last_col_letter = get_column_letter(num_cols)
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{ws.max_row}"

    @staticmethod
    def _set_print_setup(ws):
        """Sensible print defaults."""
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"

    @staticmethod
    def _write_branded_title(ws, title: str, num_cols: int, row: int = 1):
        """Write a branded title bar spanning all columns."""
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=num_cols,
        )
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = ExcelExporter.TITLE_FONT
        cell.fill = PatternFill(start_color=ExcelExporter._INDIGO_MID, end_color=ExcelExporter._INDIGO_MID, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 28

    @staticmethod
    def _write_subtitle_row(ws, text: str, num_cols: int, row: int):
        """Write a subtitle / section bar."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1)
        cell.value = text
        cell.font = ExcelExporter.SUBTITLE_FONT
        cell.fill = ExcelExporter.SUBTITLE_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def _write_info_row(ws, label: str, value, row: int, value_fmt: Optional[str] = None):
        """Write a label–value pair in columns A and B."""
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        v_cell = ws.cell(row=row, column=2, value=value)
        v_cell.font = ExcelExporter.BODY_FONT
        if value_fmt:
            v_cell.number_format = value_fmt

    @staticmethod
    def _get_merchant_for_transaction(transaction_id: int, db: Session) -> str:
        merchant_record = db.query(TransactionMerchant).filter(
            TransactionMerchant.transaction_id == transaction_id
        ).first()
        return merchant_record.merchant if merchant_record and merchant_record.merchant else ""

    @staticmethod
    def _get_all_merchants(session_id: str, db: Session, client_id: Optional[int] = None) -> Dict[str, Dict[str, Any]]:
        query = db.query(Transaction)
        if session_id:
            query = query.filter(Transaction.session_id == session_id)
        elif client_id:
            query = query.filter(Transaction.client_id == client_id)
        else:
            raise ValueError("Either session_id or client_id must be provided")
        transactions = query.all()

        merchants: Dict[str, Dict[str, Any]] = {}
        for txn in transactions:
            merchant = ExcelExporter._get_merchant_for_transaction(txn.id, db)
            if not merchant:
                merchant = "Unattributed"
            if merchant not in merchants:
                merchants[merchant] = {"total_amount": 0.0, "count": 0, "categories": {}}
            merchants[merchant]["total_amount"] += abs(txn.amount)
            merchants[merchant]["count"] += 1
            cat = txn.category or "Uncategorized"
            merchants[merchant]["categories"][cat] = merchants[merchant]["categories"].get(cat, 0) + 1
        return merchants

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        import re
        sanitized = re.sub(r'[\:\\/\?\*\[\]]', '', name)
        return sanitized[:31]

    @staticmethod
    def _build_summary_sheet(
        ws,
        transactions: list,
        categories_map: Dict[str, list],
        include_vat: bool = False,
    ):
        """Write the Category × Month pivot summary sheet.

        Layout (top → bottom):
          Title bar
          Column header row  (Category | month1 | … | Total)
          ── INCOME section banner ──
          income category rows
          VAT on Income row  (only when include_vat)
          Income Total row
          ── EXPENSES section banner ──
          expense category rows
          VAT on Expenses row  (only when include_vat)
          Expenses Total row
          ── NET row ──  (Income Total − |Expenses Total| per month)

        VAT is split by transaction sign:
          amount >= 0  →  VAT on Income
          amount <  0  →  VAT on Expenses
        Both VAT rows are included in their respective section subtotals.
        """
        # ── Collect months ────────────────────────────────────────────
        all_months_set = set()
        for t in transactions:
            all_months_set.add(t.date.strftime("%Y-%m"))
        all_months = sorted(all_months_set)

        num_data_cols = len(all_months)
        num_cols      = 1 + num_data_cols + 1  # label | months | Total
        total_col     = num_cols
        first_month_col = 2

        # ── Column widths ─────────────────────────────────────────────
        ws.column_dimensions[get_column_letter(1)].width = 26
        for ci in range(first_month_col, total_col + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 15

        # ── Title bar ─────────────────────────────────────────────────
        ExcelExporter._write_branded_title(
            ws,
            "Category Summary  –  " + ("Excl. VAT amounts" if include_vat else "Net amounts"),
            num_cols,
            row=1,
        )

        # ── Header row ────────────────────────────────────────────────
        HDR = 3
        ws.cell(row=HDR, column=1, value="Category")
        for mi, month in enumerate(all_months):
            ws.cell(row=HDR, column=first_month_col + mi, value=month)
        ws.cell(row=HDR, column=total_col, value="Total")
        ExcelExporter._style_header_row(ws, HDR, num_cols)
        for ci in range(first_month_col, total_col + 1):
            ws.cell(row=HDR, column=ci).alignment = Alignment(horizontal="right", vertical="center")

        # ── Build pivot data ───────────────────────────────────────────
        # inc_vat_per_month: VAT from amount >= 0 transactions
        # exp_vat_per_month: VAT from amount <  0 transactions
        pivot: Dict[str, Dict[str, float]] = {}
        inc_vat_per_month: Dict[str, float] = {m: 0.0 for m in all_months}
        exp_vat_per_month: Dict[str, float] = {m: 0.0 for m in all_months}

        for cat, txns in categories_map.items():
            pivot[cat] = {m: 0.0 for m in all_months}
            for t in txns:
                mk = t.date.strftime("%Y-%m")
                if include_vat:
                    vat_amt = t.vat_amount or 0.0
                    pivot[cat][mk] += t.amount - vat_amt
                    if t.amount >= 0:
                        inc_vat_per_month[mk] += vat_amt
                    else:
                        exp_vat_per_month[mk] += vat_amt
                else:
                    pivot[cat][mk] += t.amount

        # ── Classify categories into Income / Expenses ─────────────────
        # A category is Income if its overall net sum >= 0, Expenses otherwise.
        income_cats  = sorted(c for c in categories_map if sum(pivot[c].values()) >= 0)
        expense_cats = sorted(c for c in categories_map if sum(pivot[c].values()) < 0)

        # ── Style helpers ──────────────────────────────────────────────
        GREEN_FONT  = Font(name="Calibri", size=10, color=ExcelExporter._GREEN_DARK)
        RED_FONT    = Font(name="Calibri", size=10, color=ExcelExporter._RED_DARK)
        VAT_FILL    = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        VAT_FONT    = Font(name="Calibri", bold=True, size=10, color="92400E")
        INC_BANNER_FILL = PatternFill(start_color=ExcelExporter._GREEN_BG, end_color=ExcelExporter._GREEN_BG, fill_type="solid")
        INC_BANNER_FONT = Font(name="Calibri", bold=True, size=10, color=ExcelExporter._GREEN_DARK)
        EXP_BANNER_FILL = PatternFill(start_color=ExcelExporter._RED_BG, end_color=ExcelExporter._RED_BG, fill_type="solid")
        EXP_BANNER_FONT = Font(name="Calibri", bold=True, size=10, color=ExcelExporter._RED_DARK)
        NET_FILL        = PatternFill(start_color=ExcelExporter._INDIGO_DARK, end_color=ExcelExporter._INDIGO_DARK, fill_type="solid")
        NET_FONT        = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        INC_TOTAL_FNT   = Font(name="Calibri", bold=True, size=10, color=ExcelExporter._GREEN_DARK)
        EXP_TOTAL_FNT   = Font(name="Calibri", bold=True, size=10, color=ExcelExporter._RED_DARK)

        data_start_row = HDR + 1
        row = data_start_row

        # Accumulators for column-level totals
        inc_col: Dict[str, float] = {m: 0.0 for m in all_months}
        exp_col: Dict[str, float] = {m: 0.0 for m in all_months}

        # ── Helper: write one category data row ────────────────────────
        def _write_cat_row(cat: str, stripe: bool):
            nonlocal row
            lbl = ws.cell(row=row, column=1, value=cat)
            lbl.font = ExcelExporter.BODY_FONT
            lbl.border = ExcelExporter.BORDER
            row_total = 0.0
            for mi, month in enumerate(all_months):
                val = pivot[cat].get(month, 0.0)
                row_total += val
                ci = first_month_col + mi
                c = ws.cell(row=row, column=ci, value=val if val != 0.0 else None)
                c.number_format = ExcelExporter.CURRENCY_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.border = ExcelExporter.BORDER
                if val != 0.0:
                    c.font = GREEN_FONT if val >= 0 else RED_FONT
            rt = ws.cell(row=row, column=total_col, value=row_total)
            rt.number_format = ExcelExporter.CURRENCY_FMT
            rt.alignment = Alignment(horizontal="right", vertical="center")
            rt.fill = ExcelExporter.TOTAL_FILL
            rt.font = Font(name="Calibri", bold=True, size=10,
                           color=ExcelExporter._GREEN_DARK if row_total >= 0 else ExcelExporter._RED_DARK)
            rt.border = ExcelExporter.BORDER
            if stripe:
                for ci in range(1, num_cols + 1):
                    cell = ws.cell(row=row, column=ci)
                    if not cell.fill or cell.fill.fill_type == "none":
                        cell.fill = ExcelExporter.STRIPE_FILL
            row += 1

        # ── Helper: write a section banner ────────────────────────────
        def _write_banner(label: str, fill: PatternFill, font: Font):
            nonlocal row
            for ci in range(1, num_cols + 1):
                c = ws.cell(row=row, column=ci)
                c.fill = fill
                c.border = ExcelExporter.BORDER
                c.font = font
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
            row += 1

        # ── Helper: write a subtotal row (Income Total / Expenses Total) ─
        def _write_subtotal_row(label: str, col_sums: Dict[str, float], font: Font):
            nonlocal row
            row_grand = 0.0
            lbl = ws.cell(row=row, column=1, value=label)
            lbl.font = font
            lbl.fill = ExcelExporter.TOTAL_FILL
            lbl.border = ExcelExporter.BORDER
            lbl.alignment = Alignment(horizontal="left", vertical="center")
            for mi, month in enumerate(all_months):
                cv = col_sums.get(month, 0.0)
                row_grand += cv
                ci = first_month_col + mi
                c = ws.cell(row=row, column=ci, value=cv if cv != 0.0 else None)
                c.number_format = ExcelExporter.CURRENCY_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = ExcelExporter.TOTAL_FILL
                c.font = font
                c.border = ExcelExporter.BORDER
            gc = ws.cell(row=row, column=total_col, value=row_grand)
            gc.number_format = ExcelExporter.CURRENCY_FMT
            gc.alignment = Alignment(horizontal="right", vertical="center")
            gc.fill = ExcelExporter.TOTAL_FILL
            gc.font = font
            gc.border = ExcelExporter.BORDER
            row += 1
            return row_grand

        # ══════════════════════ INCOME SECTION ═══════════════════════
        _write_banner("INCOME", INC_BANNER_FILL, INC_BANNER_FONT)

        for i, cat in enumerate(income_cats):
            for m in all_months:
                inc_col[m] += pivot[cat].get(m, 0.0)
            _write_cat_row(cat, stripe=(i % 2 == 1))

        # VAT on Income row (before Income Total)
        if include_vat:
            inc_vat_total = 0.0
            lbl = ws.cell(row=row, column=1, value="VAT on Income")
            lbl.font = VAT_FONT
            lbl.fill = VAT_FILL
            lbl.border = ExcelExporter.BORDER
            for mi, month in enumerate(all_months):
                vv = inc_vat_per_month.get(month, 0.0)
                inc_vat_total += vv
                inc_col[month] += vv   # include in Income Total
                ci = first_month_col + mi
                c = ws.cell(row=row, column=ci, value=vv if vv != 0.0 else None)
                c.number_format = ExcelExporter.CURRENCY_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = VAT_FILL
                c.font = VAT_FONT
                c.border = ExcelExporter.BORDER
            vt = ws.cell(row=row, column=total_col, value=inc_vat_total)
            vt.number_format = ExcelExporter.CURRENCY_FMT
            vt.alignment = Alignment(horizontal="right", vertical="center")
            vt.fill = VAT_FILL
            vt.font = VAT_FONT
            vt.border = ExcelExporter.BORDER
            row += 1

        inc_grand = _write_subtotal_row("Income Total", inc_col, INC_TOTAL_FNT)

        # ══════════════════════ EXPENSES SECTION ═════════════════════
        _write_banner("EXPENSES", EXP_BANNER_FILL, EXP_BANNER_FONT)

        for i, cat in enumerate(expense_cats):
            for m in all_months:
                exp_col[m] += pivot[cat].get(m, 0.0)
            _write_cat_row(cat, stripe=(i % 2 == 1))

        # VAT on Expenses row (before Expenses Total)
        if include_vat:
            exp_vat_total = 0.0
            lbl = ws.cell(row=row, column=1, value="VAT on Expenses")
            lbl.font = VAT_FONT
            lbl.fill = VAT_FILL
            lbl.border = ExcelExporter.BORDER
            for mi, month in enumerate(all_months):
                vv = exp_vat_per_month.get(month, 0.0)
                exp_vat_total += vv
                exp_col[month] += vv   # include in Expenses Total
                ci = first_month_col + mi
                c = ws.cell(row=row, column=ci, value=vv if vv != 0.0 else None)
                c.number_format = ExcelExporter.CURRENCY_FMT
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = VAT_FILL
                c.font = VAT_FONT
                c.border = ExcelExporter.BORDER
            vt = ws.cell(row=row, column=total_col, value=exp_vat_total)
            vt.number_format = ExcelExporter.CURRENCY_FMT
            vt.alignment = Alignment(horizontal="right", vertical="center")
            vt.fill = VAT_FILL
            vt.font = VAT_FONT
            vt.border = ExcelExporter.BORDER
            row += 1

        exp_grand = _write_subtotal_row("Expenses Total", exp_col, EXP_TOTAL_FNT)

        # ══════════════════════ NET ROW ═══════════════════════════════
        net_grand = 0.0
        lbl = ws.cell(row=row, column=1, value="NET")
        lbl.font = NET_FONT
        lbl.fill = NET_FILL
        lbl.border = ExcelExporter.BORDER
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        for mi, month in enumerate(all_months):
            net_val = inc_col.get(month, 0.0) + exp_col.get(month, 0.0)
            net_grand += net_val
            ci = first_month_col + mi
            c = ws.cell(row=row, column=ci, value=net_val if net_val != 0.0 else None)
            c.number_format = ExcelExporter.CURRENCY_FMT
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = NET_FILL
            c.font = NET_FONT
            c.border = ExcelExporter.BORDER
        gc = ws.cell(row=row, column=total_col, value=net_grand)
        gc.number_format = ExcelExporter.CURRENCY_FMT
        gc.alignment = Alignment(horizontal="right", vertical="center")
        gc.fill = NET_FILL
        gc.font = NET_FONT
        gc.border = ExcelExporter.BORDER
        row += 1

        # ── Freeze panes (keep header + category column visible) ───────
        ws.freeze_panes = ws.cell(row=data_start_row, column=first_month_col)
        ExcelExporter._set_print_setup(ws)

    # ══════════════════════════════════════════════════════════
    #  EXPORT: TRANSACTIONS
    # ══════════════════════════════════════════════════════════
    @staticmethod
    def export_transactions(session_id: Optional[str], db: Session, client_id: Optional[int] = None, include_vat: bool = False, client_name: Optional[str] = None) -> BytesIO:
        """Export all transactions – clean, professional format."""
        query = db.query(Transaction)
        if session_id:
            query = query.filter(Transaction.session_id == session_id)
        elif client_id:
            query = query.filter(Transaction.client_id == client_id)
        else:
            raise ValueError("Either session_id or client_id must be provided")

        transactions = query.order_by(Transaction.date).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        if include_vat:
            NUM_COLS = 6
            col_widths = {"A": 14, "B": 42, "C": 18, "D": 16, "E": 18, "F": 20}
            headers = ["Date", "Description", "Amount (Incl VAT)", "VAT Amount", "Amount (Excl VAT)", "Category"]
        else:
            NUM_COLS = 4
            col_widths = {"A": 14, "B": 42, "C": 18, "D": 20}
            headers = ["Date", "Description", "Amount", "Category"]

        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w

        # Title bar
        ExcelExporter._write_branded_title(ws, "Transaction Export", NUM_COLS, row=1)

        # Info rows
        if transactions:
            date_range = f"{transactions[0].date.isoformat()}  to  {transactions[-1].date.isoformat()}"
            ExcelExporter._write_info_row(ws, "Period:", date_range, row=2)
        ExcelExporter._write_info_row(ws, "Transactions:", len(transactions), row=3)
        if client_name:
            ExcelExporter._write_info_row(ws, "Client:", client_name, row=4)

        # Header row
        HEADER_ROW = 5
        for col, header in enumerate(headers, 1):
            ws.cell(row=HEADER_ROW, column=col, value=header)
        ExcelExporter._style_header_row(ws, HEADER_ROW, NUM_COLS)

        # Data rows
        for i, txn in enumerate(transactions):
            r = HEADER_ROW + 1 + i
            ws.cell(row=r, column=1, value=txn.date.isoformat())
            ws.cell(row=r, column=2, value=txn.description)

            if include_vat:
                vat_amount = txn.vat_amount if txn.vat_amount is not None else 0.0
                amount_incl = txn.amount
                amount_excl = amount_incl - vat_amount
                c3 = ws.cell(row=r, column=3, value=amount_incl)
                c3.number_format = ExcelExporter.CURRENCY_FMT
                c4 = ws.cell(row=r, column=4, value=vat_amount)
                c4.number_format = ExcelExporter.CURRENCY_FMT
                c5 = ws.cell(row=r, column=5, value=amount_excl)
                c5.number_format = ExcelExporter.CURRENCY_FMT
                ws.cell(row=r, column=6, value=txn.category or "Uncategorized")
            else:
                amt_cell = ws.cell(row=r, column=3, value=txn.amount)
                ws.cell(row=r, column=4, value=txn.category or "Uncategorized")
                ExcelExporter._style_amount_cell(amt_cell, txn.amount)

            ExcelExporter._style_data_row(ws, r, NUM_COLS, is_stripe=(i % 2 == 1))

        # Totals row
        if transactions:
            total_r = HEADER_ROW + 1 + len(transactions)
            ws.cell(row=total_r, column=1, value="TOTAL")
            ws.cell(row=total_r, column=3, value=f"=SUM(C{HEADER_ROW+1}:C{total_r-1})")
            ws.cell(row=total_r, column=3).number_format = ExcelExporter.CURRENCY_FMT
            if include_vat:
                ws.cell(row=total_r, column=4, value=f"=SUM(D{HEADER_ROW+1}:D{total_r-1})")
                ws.cell(row=total_r, column=4).number_format = ExcelExporter.CURRENCY_FMT
                ws.cell(row=total_r, column=5, value=f"=SUM(E{HEADER_ROW+1}:E{total_r-1})")
                ws.cell(row=total_r, column=5).number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_total_row(ws, total_r, NUM_COLS)

        ExcelExporter._freeze_and_filter(ws, HEADER_ROW, NUM_COLS)
        ExcelExporter._set_print_setup(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════
    #  EXPORT: FOR ACCOUNTANT
    # ══════════════════════════════════════════════════════════
    @staticmethod
    def export_for_accountant(session_id: str, db: Session, client_id: Optional[int] = None, include_vat: bool = False, client_name: Optional[str] = None) -> BytesIO:
        """Comprehensive multi-sheet report for accountants."""
        query = db.query(Transaction)
        if session_id:
            query = query.filter(Transaction.session_id == session_id)
        elif client_id:
            query = query.filter(Transaction.client_id == client_id)
        else:
            raise ValueError("Either session_id or client_id must be provided")
        transactions = query.order_by(Transaction.date).all()

        if not transactions:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "No Data"
            ws["A1"] = "No transactions found for this session"
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        merchants = ExcelExporter._get_all_merchants(session_id, db, client_id)

        wb = openpyxl.Workbook()

        # ── Sheet 1: Executive Summary ────────────────────────
        ws_s = wb.active
        ws_s.title = "Executive Summary"
        ws_s.column_dimensions["A"].width = 28
        ws_s.column_dimensions["B"].width = 20

        ExcelExporter._write_branded_title(ws_s, "Executive Summary", 2, row=1)

        date_range = f"{transactions[0].date.isoformat()}  to  {transactions[-1].date.isoformat()}"
        ExcelExporter._write_info_row(ws_s, "Period:", date_range, row=3)
        ExcelExporter._write_info_row(ws_s, "Generated:", date.today().isoformat(), row=4)
        if client_name:
            ExcelExporter._write_info_row(ws_s, "Client:", client_name, row=5)

        # Key Metrics
        total_income = sum(t.amount for t in transactions if t.amount >= 0)
        total_expenses = sum(abs(t.amount) for t in transactions if t.amount < 0)
        net_balance = total_income - total_expenses

        # VAT totals (when applicable)
        total_vat = sum((t.vat_amount or 0.0) for t in transactions) if include_vat else 0.0
        total_excl_vat = sum(t.amount - (t.vat_amount or 0.0) for t in transactions) if include_vat else 0.0

        row = 6
        ExcelExporter._write_subtitle_row(ws_s, "KEY METRICS", 2, row)
        row += 1
        metrics = [
            ("Total Income", total_income),
            ("Total Expenses", total_expenses),
            ("Net Balance", net_balance),
            ("Transaction Count", len(transactions)),
            ("Unique Merchants", len(merchants)),
        ]
        if include_vat:
            metrics.append(("Total VAT", total_vat))
            metrics.append(("Total Excl. VAT", total_excl_vat))
        for label, val in metrics:
            ExcelExporter._write_info_row(ws_s, label, val, row, ExcelExporter.CURRENCY_FMT if isinstance(val, float) else None)
            if isinstance(val, float):
                ExcelExporter._style_amount_cell(ws_s.cell(row=row, column=2), val)
            row += 1

        # Income by Category
        row += 1
        ExcelExporter._write_subtitle_row(ws_s, "INCOME BY CATEGORY", 2, row)
        row += 1
        for col, h in enumerate(["Category", "Amount"], 1):
            ws_s.cell(row=row, column=col, value=h)
        ExcelExporter._style_header_row(ws_s, row, 2)
        row += 1

        income_cats: Dict[str, float] = {}
        for t in transactions:
            if t.amount >= 0:
                cat = t.category or "Uncategorized"
                income_cats[cat] = income_cats.get(cat, 0) + t.amount
        sorted_income = sorted(income_cats.items(), key=lambda x: x[1], reverse=True)
        inc_start = row
        for i, (cat, amount) in enumerate(sorted_income):
            ws_s.cell(row=row, column=1, value=cat)
            c = ws_s.cell(row=row, column=2, value=amount)
            c.number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_data_row(ws_s, row, 2, is_stripe=(i % 2 == 1))
            row += 1
        if sorted_income:
            ws_s.cell(row=row, column=1, value="Income Subtotal")
            ws_s.cell(row=row, column=2, value=f"=SUM(B{inc_start}:B{row-1})")
            ws_s.cell(row=row, column=2).number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_total_row(ws_s, row, 2)
            row += 2

        # Expenses by Category
        ExcelExporter._write_subtitle_row(ws_s, "EXPENSES BY CATEGORY", 2, row)
        row += 1
        for col, h in enumerate(["Category", "Amount"], 1):
            ws_s.cell(row=row, column=col, value=h)
        ExcelExporter._style_header_row(ws_s, row, 2)
        row += 1

        expense_cats: Dict[str, float] = {}
        for t in transactions:
            if t.amount < 0:
                cat = t.category or "Uncategorized"
                expense_cats[cat] = expense_cats.get(cat, 0) + abs(t.amount)
        sorted_expenses = sorted(expense_cats.items(), key=lambda x: x[1], reverse=True)
        exp_start = row
        for i, (cat, amount) in enumerate(sorted_expenses):
            ws_s.cell(row=row, column=1, value=cat)
            c = ws_s.cell(row=row, column=2, value=amount)
            c.number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_data_row(ws_s, row, 2, is_stripe=(i % 2 == 1))
            row += 1
        if sorted_expenses:
            ws_s.cell(row=row, column=1, value="Expense Subtotal")
            ws_s.cell(row=row, column=2, value=f"=SUM(B{exp_start}:B{row-1})")
            ws_s.cell(row=row, column=2).number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_total_row(ws_s, row, 2)
            row += 2

        # Net Balance highlight
        ExcelExporter._write_subtitle_row(ws_s, "NET BALANCE", 2, row)
        row += 1
        net_fill_color = ExcelExporter._GREEN_BG if net_balance >= 0 else ExcelExporter._RED_BG
        net_font_color = ExcelExporter._GREEN_DARK if net_balance >= 0 else ExcelExporter._RED_DARK
        ws_s.cell(row=row, column=1, value="Net Balance")
        ws_s.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12, color=net_font_color)
        ws_s.cell(row=row, column=1).fill = PatternFill(start_color=net_fill_color, end_color=net_fill_color, fill_type="solid")
        nb_cell = ws_s.cell(row=row, column=2, value=net_balance)
        nb_cell.font = Font(name="Calibri", bold=True, size=12, color=net_font_color)
        nb_cell.fill = PatternFill(start_color=net_fill_color, end_color=net_fill_color, fill_type="solid")
        nb_cell.number_format = ExcelExporter.CURRENCY_FMT

        ExcelExporter._set_print_setup(ws_s)

        # ── Sheet 2: Merchant Analysis ────────────────────────
        ws_m = wb.create_sheet("Merchant Analysis")
        NUM_M = 4
        for letter, w in {"A": 32, "B": 18, "C": 18, "D": 22}.items():
            ws_m.column_dimensions[letter].width = w

        ExcelExporter._write_branded_title(ws_m, "Merchant Analysis", NUM_M, row=1)

        HEADER_R = 3
        for col, h in enumerate(["Merchant", "Transactions", "Total Amount", "Avg Per Transaction"], 1):
            ws_m.cell(row=HEADER_R, column=col, value=h)
        ExcelExporter._style_header_row(ws_m, HEADER_R, NUM_M)

        sorted_merchants = sorted(merchants.items(), key=lambda x: x[1]["total_amount"], reverse=True)
        for i, (name, data) in enumerate(sorted_merchants):
            r = HEADER_R + 1 + i
            ws_m.cell(row=r, column=1, value=name)
            ws_m.cell(row=r, column=2, value=data["count"])
            amt_c = ws_m.cell(row=r, column=3, value=data["total_amount"])
            amt_c.number_format = ExcelExporter.CURRENCY_FMT
            avg = data["total_amount"] / data["count"] if data["count"] else 0
            avg_c = ws_m.cell(row=r, column=4, value=avg)
            avg_c.number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_data_row(ws_m, r, NUM_M, is_stripe=(i % 2 == 1))

        ExcelExporter._freeze_and_filter(ws_m, HEADER_R, NUM_M)
        ExcelExporter._set_print_setup(ws_m)

        # ── Sheet 3: Detailed Transactions ────────────────────
        ws_d = wb.create_sheet("Transactions")

        if include_vat:
            NUM_D = 7
            d_col_widths = {"A": 14, "B": 42, "C": 28, "D": 18, "E": 16, "F": 18, "G": 20}
            d_headers = ["Date", "Description", "Merchant", "Amount (Incl VAT)", "VAT Amount", "Amount (Excl VAT)", "Category"]
        else:
            NUM_D = 5
            d_col_widths = {"A": 14, "B": 42, "C": 28, "D": 18, "E": 20}
            d_headers = ["Date", "Description", "Merchant", "Amount", "Category"]

        for letter, w in d_col_widths.items():
            ws_d.column_dimensions[letter].width = w

        ExcelExporter._write_branded_title(ws_d, "Detailed Transactions", NUM_D, row=1)

        D_HEADER = 3
        for col, h in enumerate(d_headers, 1):
            ws_d.cell(row=D_HEADER, column=col, value=h)
        ExcelExporter._style_header_row(ws_d, D_HEADER, NUM_D)

        for i, txn in enumerate(transactions):
            r = D_HEADER + 1 + i
            merchant = ExcelExporter._get_merchant_for_transaction(txn.id, db) or ""
            ws_d.cell(row=r, column=1, value=txn.date.isoformat())
            ws_d.cell(row=r, column=2, value=txn.description)
            ws_d.cell(row=r, column=3, value=merchant)

            if include_vat:
                vat_amt = txn.vat_amount if txn.vat_amount is not None else 0.0
                amt_incl = txn.amount
                amt_excl = amt_incl - vat_amt
                c4 = ws_d.cell(row=r, column=4, value=amt_incl)
                c4.number_format = ExcelExporter.CURRENCY_FMT
                c5 = ws_d.cell(row=r, column=5, value=vat_amt)
                c5.number_format = ExcelExporter.CURRENCY_FMT
                c6 = ws_d.cell(row=r, column=6, value=amt_excl)
                c6.number_format = ExcelExporter.CURRENCY_FMT
                ws_d.cell(row=r, column=7, value=txn.category or "Uncategorized")
            else:
                amt_c = ws_d.cell(row=r, column=4, value=txn.amount)
                ws_d.cell(row=r, column=5, value=txn.category or "Uncategorized")
                ExcelExporter._style_amount_cell(amt_c, txn.amount)

            ExcelExporter._style_data_row(ws_d, r, NUM_D, is_stripe=(i % 2 == 1))

        if transactions:
            total_r = D_HEADER + 1 + len(transactions)
            ws_d.cell(row=total_r, column=1, value="TOTAL")
            ws_d.cell(row=total_r, column=4, value=f"=SUM(D{D_HEADER+1}:D{total_r-1})")
            ws_d.cell(row=total_r, column=4).number_format = ExcelExporter.CURRENCY_FMT
            if include_vat:
                ws_d.cell(row=total_r, column=5, value=f"=SUM(E{D_HEADER+1}:E{total_r-1})")
                ws_d.cell(row=total_r, column=5).number_format = ExcelExporter.CURRENCY_FMT
                ws_d.cell(row=total_r, column=6, value=f"=SUM(F{D_HEADER+1}:F{total_r-1})")
                ws_d.cell(row=total_r, column=6).number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_total_row(ws_d, total_r, NUM_D)

        ExcelExporter._freeze_and_filter(ws_d, D_HEADER, NUM_D)
        ExcelExporter._set_print_setup(ws_d)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════
    #  EXPORT: UNUSUAL TRANSACTIONS REPORT
    # ══════════════════════════════════════════════════════════
    _FLAG_LABELS: dict = {
        "above_threshold": "Above threshold",
        "below_threshold": "Below threshold",
        "round_amount":    "Round amount",
        "recurring":       "Recurring",
        "duplicate":       "Possible duplicate",
        "weekend":         "Weekend",
    }

    @staticmethod
    def export_unusual_transactions(
        flagged_transactions: list,
        client_name: Optional[str] = None,
    ) -> BytesIO:
        """Export unusual/flagged transactions to a styled Excel workbook."""
        NUM_COLS = 6
        col_widths = {"A": 14, "B": 40, "C": 22, "D": 18, "E": 22, "F": 42}
        headers = ["Date", "Description", "Merchant", "Amount", "Category", "Flags"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unusual Transactions"

        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w

        # ── Title bar ─────────────────────────────────────────
        ExcelExporter._write_branded_title(ws, "Unusual Transactions Report", NUM_COLS, row=1)

        # ── Info rows ─────────────────────────────────────────
        ExcelExporter._write_info_row(ws, "Flagged transactions:", len(flagged_transactions), row=2)
        if client_name:
            ExcelExporter._write_info_row(ws, "Client:", client_name, row=3)

        # ── Header row ────────────────────────────────────────
        HEADER_ROW = 4
        for col, h in enumerate(headers, 1):
            ws.cell(row=HEADER_ROW, column=col, value=h)
        ExcelExporter._style_header_row(ws, HEADER_ROW, NUM_COLS)

        # ── Data rows ─────────────────────────────────────────
        FLAG_LABELS = ExcelExporter._FLAG_LABELS
        for i, txn in enumerate(flagged_transactions):
            r = HEADER_ROW + 1 + i
            amount = txn.get("amount", 0)
            flag_keys: list = txn.get("flags", [])
            flag_text = "; ".join(FLAG_LABELS.get(f, f) for f in flag_keys)

            ws.cell(row=r, column=1, value=txn.get("date", ""))
            ws.cell(row=r, column=2, value=txn.get("description", ""))
            ws.cell(row=r, column=3, value=txn.get("merchant", "") or "")
            amt_cell = ws.cell(row=r, column=4, value=amount)
            ws.cell(row=r, column=5, value=txn.get("category") or "Uncategorized")
            ws.cell(row=r, column=6, value=flag_text)

            ExcelExporter._style_data_row(ws, r, NUM_COLS, is_stripe=(i % 2 == 1))
            ExcelExporter._style_amount_cell(amt_cell, amount)

        ExcelExporter._freeze_and_filter(ws, HEADER_ROW, NUM_COLS)
        ExcelExporter._set_print_setup(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════
    #  EXPORT: SINGLE CATEGORY (monthly sections)
    # ══════════════════════════════════════════════════════════
    @staticmethod
    def export_category_monthly(session_id: str, category: str, db: Session, client_id: Optional[int] = None, client_name: Optional[str] = None) -> BytesIO:
        query = db.query(Transaction).filter(Transaction.category == category)
        if session_id:
            query = query.filter(Transaction.session_id == session_id)
        elif client_id:
            query = query.filter(Transaction.client_id == client_id)
        else:
            raise ValueError("Either session_id or client_id must be provided")
        txns = query.order_by(Transaction.date).all()

        wb = openpyxl.Workbook()
        sheet_name = ExcelExporter._sanitize_sheet_name(category)
        ws = wb.active
        ws.title = sheet_name
        NUM_COLS = 4

        for letter, w in {"A": 14, "B": 42, "C": 18, "D": 20}.items():
            ws.column_dimensions[letter].width = w

        ExcelExporter._write_branded_title(ws, f"Category: {category}", NUM_COLS, row=1)
        if client_name:
            ExcelExporter._write_info_row(ws, "Client:", client_name, row=2)
        row = 3

        if not txns:
            ws.cell(row=row, column=1, value="No transactions for this category")
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # Group by month
        months: Dict[str, List[Any]] = {}
        for t in txns:
            m = t.date.strftime("%Y-%m")
            months.setdefault(m, []).append(t)

        month_keys = sorted(months.keys())
        all_txns_by_month = [(mk, months[mk]) for mk in month_keys]

        # Cumulative sums
        cumulative_before: Dict[str, float] = {}
        for idx, (mk, _) in enumerate(all_txns_by_month):
            cumulative_before[mk] = sum(
                pt.amount for j in range(idx) for pt in all_txns_by_month[j][1]
            )

        for mk, mtxns in all_txns_by_month:
            # Month header
            ExcelExporter._write_subtitle_row(ws, mk, NUM_COLS, row)
            row += 1

            # Opening balance
            ws.cell(row=row, column=1, value="Opening balance").font = Font(name="Calibri", italic=True, size=10)
            ob_cell = ws.cell(row=row, column=NUM_COLS, value=cumulative_before.get(mk, 0.0))
            ob_cell.number_format = ExcelExporter.CURRENCY_FMT
            ob_cell.font = Font(name="Calibri", italic=True, size=10)
            row += 1

            # Column headers
            for col, h in enumerate(["Date", "Description", "Amount", "Running Balance"], 1):
                ws.cell(row=row, column=col, value=h)
            ExcelExporter._style_header_row(ws, row, NUM_COLS)
            row += 1

            running = cumulative_before.get(mk, 0.0)
            for i, t in enumerate(mtxns):
                ws.cell(row=row, column=1, value=t.date.isoformat())
                ws.cell(row=row, column=2, value=t.description)
                amt_c = ws.cell(row=row, column=3, value=t.amount)
                running += t.amount
                bal_c = ws.cell(row=row, column=4, value=running)
                ExcelExporter._style_data_row(ws, row, NUM_COLS, is_stripe=(i % 2 == 1))
                ExcelExporter._style_amount_cell(amt_c, t.amount)
                bal_c.number_format = ExcelExporter.CURRENCY_FMT
                row += 1

            row += 1  # blank between months

        ExcelExporter._set_print_setup(ws)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════
    #  EXPORT: ALL CATEGORIES (one sheet per category)
    # ══════════════════════════════════════════════════════════
    @staticmethod
    def export_all_categories_monthly(
        session_id: str,
        db: Session,
        client_id: Optional[int] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        include_vat: bool = False,
        selected_categories: Optional[List[str]] = None,
        client_name: Optional[str] = None,
    ) -> BytesIO:
        from datetime import datetime

        query = db.query(Transaction)
        if session_id:
            query = query.filter(Transaction.session_id == session_id)
        elif client_id:
            query = query.filter(Transaction.client_id == client_id)
        else:
            raise ValueError("Either session_id or client_id must be provided")
        if date_from:
            query = query.filter(Transaction.date >= datetime.strptime(date_from, "%Y-%m-%d").date())
        if date_to:
            query = query.filter(Transaction.date <= datetime.strptime(date_to, "%Y-%m-%d").date())
        transactions = query.order_by(Transaction.date).all()

        if not transactions:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "No Data"
            ws["A1"] = "No transactions found for this session"
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # Group by category
        categories_map: Dict[str, List[Any]] = {}
        for t in transactions:
            cat_name = t.category or "Uncategorized"
            categories_map.setdefault(cat_name, []).append(t)

        if selected_categories:
            categories_map = {k: v for k, v in categories_map.items() if k in selected_categories}

        if include_vat:
            num_cols = 6
            col_widths = {"A": 14, "B": 42, "C": 18, "D": 16, "E": 18, "F": 20}
            headers = ["Date", "Description", "Amount (Incl VAT)", "VAT Amount", "Amount (Excl VAT)", "Running Balance"]
        else:
            num_cols = 4
            col_widths = {"A": 14, "B": 42, "C": 18, "D": 20}
            headers = ["Date", "Description", "Amount", "Running Balance"]

        wb = openpyxl.Workbook()

        # ── Summary sheet (pivot: categories × months) ─────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ExcelExporter._build_summary_sheet(
            ws_summary,
            transactions,
            categories_map,
            include_vat=include_vat,
        )

        for category, txns in categories_map.items():
            sheet_name = ExcelExporter._sanitize_sheet_name(category)
            # Always create new sheets — Summary already occupies wb.active
            ws = wb.create_sheet(title=sheet_name)

            for letter, w in col_widths.items():
                ws.column_dimensions[letter].width = w

            ExcelExporter._write_branded_title(ws, f"Category: {category}", num_cols, row=1)
            if client_name:
                ExcelExporter._write_info_row(ws, "Client:", client_name, row=2)

            row = 3
            data_first_row = None

            # Group by month
            months: Dict[str, List[Any]] = {}
            for t in txns:
                m = t.date.strftime("%Y-%m")
                months.setdefault(m, []).append(t)
            month_keys = sorted(months.keys())
            all_txns_by_month = [(mk, months[mk]) for mk in month_keys]

            for idx, (mk, mtxns) in enumerate(all_txns_by_month):
                prev_sum = sum(
                    pt.amount for j in range(idx) for pt in all_txns_by_month[j][1]
                )

                # Month section header
                ExcelExporter._write_subtitle_row(ws, mk, num_cols, row)
                row += 1

                # Opening balance
                ws.cell(row=row, column=1, value="Opening balance").font = Font(name="Calibri", italic=True, size=10)
                ob_cell = ws.cell(row=row, column=num_cols, value=prev_sum)
                ob_cell.number_format = ExcelExporter.CURRENCY_FMT
                ob_cell.font = Font(name="Calibri", italic=True, size=10)
                row += 1

                # Column headers
                for col, h in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=h)
                ExcelExporter._style_header_row(ws, row, num_cols)
                row += 1

                running = prev_sum
                for i, t in enumerate(mtxns):
                    ws.cell(row=row, column=1, value=t.date.isoformat())
                    ws.cell(row=row, column=2, value=t.description)

                    # mark the first data row for this sheet (used for totals)
                    if data_first_row is None:
                        data_first_row = row

                    if include_vat:
                        vat_amount = t.vat_amount if t.vat_amount is not None else 0.0
                        amount_incl_vat = t.amount
                        amount_excl_vat = amount_incl_vat - vat_amount

                        c3 = ws.cell(row=row, column=3, value=amount_incl_vat)
                        c3.number_format = ExcelExporter.CURRENCY_FMT
                        c4 = ws.cell(row=row, column=4, value=vat_amount)
                        c4.number_format = ExcelExporter.CURRENCY_FMT
                        c5 = ws.cell(row=row, column=5, value=amount_excl_vat)
                        c5.number_format = ExcelExporter.CURRENCY_FMT

                        running += t.amount
                        c6 = ws.cell(row=row, column=6, value=running)
                        c6.number_format = ExcelExporter.CURRENCY_FMT
                    else:
                        amt_c = ws.cell(row=row, column=3, value=t.amount)
                        running += t.amount
                        bal_c = ws.cell(row=row, column=4, value=running)
                        ExcelExporter._style_amount_cell(amt_c, t.amount)
                        bal_c.number_format = ExcelExporter.CURRENCY_FMT

                    ExcelExporter._style_data_row(ws, row, num_cols, is_stripe=(i % 2 == 1))
                    row += 1

                row += 1  # blank between months

            # After listing all months, write totals for the sheet
            if data_first_row is not None:
                total_label_row = row
                ws.cell(row=total_label_row, column=1, value="TOTAL")
                if include_vat:
                    # Amount (Incl VAT) is col 3, VAT Amount col 4, Amount Excl VAT col 5
                    ws.cell(row=total_label_row, column=3, value=f"=SUM(C{data_first_row}:C{row-1})")
                    ws.cell(row=total_label_row, column=3).number_format = ExcelExporter.CURRENCY_FMT
                    ws.cell(row=total_label_row, column=4, value=f"=SUM(D{data_first_row}:D{row-1})")
                    ws.cell(row=total_label_row, column=4).number_format = ExcelExporter.CURRENCY_FMT
                    ws.cell(row=total_label_row, column=5, value=f"=SUM(E{data_first_row}:E{row-1})")
                    ws.cell(row=total_label_row, column=5).number_format = ExcelExporter.CURRENCY_FMT
                else:
                    # Amount is col 3 when VAT not included
                    ws.cell(row=total_label_row, column=3, value=f"=SUM(C{data_first_row}:C{row-1})")
                    ws.cell(row=total_label_row, column=3).number_format = ExcelExporter.CURRENCY_FMT

                ExcelExporter._style_total_row(ws, total_label_row, num_cols)
                row += 1

            ExcelExporter._set_print_setup(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ══════════════════════════════════════════════════════════
    #  EXPORT: MONTHLY SUMMARY
    # ══════════════════════════════════════════════════════════
    @staticmethod
    def export_monthly_summary(summary_data: Dict[str, Any], include_vat: bool = False, transactions: Optional[List[Any]] = None, client_name: Optional[str] = None) -> BytesIO:
        wb = openpyxl.Workbook()

        # ── Sheet 1: Monthly Overview ─────────────────────────
        ws_o = wb.active
        ws_o.title = "Monthly Summary"

        if include_vat:
            NUM_O = 7
            o_headers = [
                "Month",
                "Total Income",
                "Total Expenses",
                "Net Balance",
                "VAT Output",
                "VAT Input",
                "Total Excl. VAT",
            ]
            for letter, w in {"A": 14, "B": 18, "C": 18, "D": 18, "E": 16, "F": 16, "G": 18}.items():
                ws_o.column_dimensions[letter].width = w
        else:
            NUM_O = 4
            o_headers = ["Month", "Total Income", "Total Expenses", "Net Balance"]
            for letter, w in {"A": 14, "B": 18, "C": 18, "D": 18}.items():
                ws_o.column_dimensions[letter].width = w

        ExcelExporter._write_branded_title(ws_o, "Monthly Summary", NUM_O, row=1)
        if client_name:
            ExcelExporter._write_info_row(ws_o, "Client:", client_name, row=2)

        HEADER_R = 3
        for col, h in enumerate(o_headers, 1):
            ws_o.cell(row=HEADER_R, column=col, value=h)
        ExcelExporter._style_header_row(ws_o, HEADER_R, NUM_O)

        months = summary_data.get("months", [])

        # Pre-compute per-month VAT output/input and excl totals from transactions if available
        month_output_vat: Dict[str, float] = {}
        month_input_vat: Dict[str, float] = {}
        month_excl: Dict[str, float] = {}
        if include_vat and transactions:
            for t in transactions:
                mk = t.date.strftime("%Y-%m")
                vat = t.vat_amount if t.vat_amount is not None else 0.0
                amt = t.amount if t.amount is not None else 0.0
                # classify VAT: positive amounts -> output, negative amounts -> input
                if amt >= 0:
                    month_output_vat[mk] = month_output_vat.get(mk, 0.0) + abs(vat)
                else:
                    month_input_vat[mk] = month_input_vat.get(mk, 0.0) + abs(vat)
                month_excl[mk] = month_excl.get(mk, 0.0) + (amt - vat)

        for i, month in enumerate(months):
            r = HEADER_R + 1 + i
            ws_o.cell(row=r, column=1, value=month["month"])
            for col_idx, key in enumerate(["total_income", "total_expenses", "net_balance"], 2):
                c = ws_o.cell(row=r, column=col_idx, value=month[key])
                c.number_format = ExcelExporter.CURRENCY_FMT
                if key == "net_balance":
                    ExcelExporter._style_amount_cell(c, month[key])
            if include_vat:
                mk = month["month"]
                c5 = ws_o.cell(row=r, column=5, value=month_output_vat.get(mk, 0.0))
                c5.number_format = ExcelExporter.CURRENCY_FMT
                c6 = ws_o.cell(row=r, column=6, value=month_input_vat.get(mk, 0.0))
                c6.number_format = ExcelExporter.CURRENCY_FMT
                c7 = ws_o.cell(row=r, column=7, value=month_excl.get(mk, 0.0))
                c7.number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_data_row(ws_o, r, NUM_O, is_stripe=(i % 2 == 1))

        # Overall totals
        if months:
            total_r = HEADER_R + 1 + len(months)
            overall = summary_data.get("overall", {})
            ws_o.cell(row=total_r, column=1, value="OVERALL")
            for col_idx, key in enumerate(["total_income", "total_expenses", "net_balance"], 2):
                c = ws_o.cell(row=total_r, column=col_idx, value=overall.get(key, 0))
                c.number_format = ExcelExporter.CURRENCY_FMT
            if include_vat:
                total_output_vat = sum(month_output_vat.values())
                total_input_vat = sum(month_input_vat.values())
                total_excl_val = sum(month_excl.values())
                ws_o.cell(row=total_r, column=5, value=total_output_vat)
                ws_o.cell(row=total_r, column=5).number_format = ExcelExporter.CURRENCY_FMT
                ws_o.cell(row=total_r, column=6, value=total_input_vat)
                ws_o.cell(row=total_r, column=6).number_format = ExcelExporter.CURRENCY_FMT
                ws_o.cell(row=total_r, column=7, value=total_excl_val)
                ws_o.cell(row=total_r, column=7).number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_total_row(ws_o, total_r, NUM_O)

        ExcelExporter._freeze_and_filter(ws_o, HEADER_R, NUM_O)
        ExcelExporter._set_print_setup(ws_o)

        # ── Sheet 2: Category Breakdown ───────────────────────
        ws_c = wb.create_sheet("Category Breakdown")

        if include_vat:
            NUM_C = 5
            c_headers = ["Category", "Amount", "VAT Output", "VAT Input", "Amount Excl. VAT"]
            for letter, w in {"A": 24, "B": 18, "C": 16, "D": 16, "E": 20}.items():
                ws_c.column_dimensions[letter].width = w
        else:
            NUM_C = 2
            c_headers = ["Category", "Total Amount"]
            ws_c.column_dimensions["A"].width = 24
            ws_c.column_dimensions["B"].width = 18

        ExcelExporter._write_branded_title(ws_c, "Category Breakdown", NUM_C, row=1)

        C_HEADER = 3
        for col, h in enumerate(c_headers, 1):
            ws_c.cell(row=C_HEADER, column=col, value=h)
        ExcelExporter._style_header_row(ws_c, C_HEADER, NUM_C)

        all_categories: Dict[str, float] = {}
        cat_output_vat: Dict[str, float] = {}
        cat_input_vat: Dict[str, float] = {}
        cat_excl: Dict[str, float] = {}

        # Prefer computing signed per-category totals from the raw transactions when available
        if transactions:
            for t in transactions:
                cat = t.category or "Uncategorised"
                amt = t.amount if t.amount is not None else 0.0
                vat = t.vat_amount if t.vat_amount is not None else 0.0

                # Signed gross amount per category
                all_categories[cat] = all_categories.get(cat, 0.0) + amt

                # Classify VAT: income (amt >= 0) -> output VAT, expense (amt < 0) -> input VAT
                if amt >= 0:
                    cat_output_vat[cat] = cat_output_vat.get(cat, 0.0) + abs(vat)
                else:
                    cat_input_vat[cat] = cat_input_vat.get(cat, 0.0) + abs(vat)

                # Amount excluding VAT (signed)
                cat_excl[cat] = cat_excl.get(cat, 0.0) + (amt - vat)
        else:
            # Fallback to month-aggregated values (these are absolute amounts)
            for month in months:
                for category, amount in month.get("categories", {}).items():
                    all_categories[category] = all_categories.get(category, 0) + amount

        # Sort categories by absolute total (largest contributors first)
        sorted_cats = sorted(all_categories.items(), key=lambda x: abs(x[1]), reverse=True)

        for i, (category, amount) in enumerate(sorted_cats):
            r = C_HEADER + 1 + i
            ws_c.cell(row=r, column=1, value=category)
            # Amount (signed gross). Prefer transaction-sourced amount when available
            if transactions:
                display_amount = all_categories.get(category, 0.0)
            else:
                display_amount = amount
            c = ws_c.cell(row=r, column=2, value=display_amount)
            c.number_format = ExcelExporter.CURRENCY_FMT
            if include_vat:
                cv_out = ws_c.cell(row=r, column=3, value=cat_output_vat.get(category, 0.0))
                cv_out.number_format = ExcelExporter.CURRENCY_FMT
                cv_in = ws_c.cell(row=r, column=4, value=cat_input_vat.get(category, 0.0))
                cv_in.number_format = ExcelExporter.CURRENCY_FMT
                ce = ws_c.cell(row=r, column=5, value=cat_excl.get(category, 0.0))
                ce.number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_data_row(ws_c, r, NUM_C, is_stripe=(i % 2 == 1))

        if all_categories:
            total_r = C_HEADER + 1 + len(all_categories)
            ws_c.cell(row=total_r, column=1, value="TOTAL")
            ws_c.cell(row=total_r, column=2, value=f"=SUM(B{C_HEADER+1}:B{total_r-1})")
            ws_c.cell(row=total_r, column=2).number_format = ExcelExporter.CURRENCY_FMT
            if include_vat:
                ws_c.cell(row=total_r, column=3, value=f"=SUM(C{C_HEADER+1}:C{total_r-1})")
                ws_c.cell(row=total_r, column=3).number_format = ExcelExporter.CURRENCY_FMT
                ws_c.cell(row=total_r, column=4, value=f"=SUM(D{C_HEADER+1}:D{total_r-1})")
                ws_c.cell(row=total_r, column=4).number_format = ExcelExporter.CURRENCY_FMT
                ws_c.cell(row=total_r, column=5, value=f"=SUM(E{C_HEADER+1}:E{total_r-1})")
                ws_c.cell(row=total_r, column=5).number_format = ExcelExporter.CURRENCY_FMT
            ExcelExporter._style_total_row(ws_c, total_r, NUM_C)

        ExcelExporter._freeze_and_filter(ws_c, C_HEADER, NUM_C)
        ExcelExporter._set_print_setup(ws_c)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
