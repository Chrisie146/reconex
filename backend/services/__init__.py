"""
Export Service
Generates Excel files for transactions and monthly summaries
Accountant-ready format
"""

import os
from io import BytesIO
from datetime import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from models import Transaction


class ExcelExporter:
    """Generate accountant-ready Excel exports"""
    
    # Style definitions
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    def export_transactions(session_id: str, db: Session) -> BytesIO:
        """
        Export all transactions to Excel
        Clean, professional format suitable for accountants
        """
        
        # Fetch transactions
        transactions = db.query(Transaction).filter(
            Transaction.session_id == session_id
        ).order_by(Transaction.date).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Headers
        headers = ["Date", "Description", "Amount", "Category"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = ExcelExporter.BORDER
        
        # Data rows
        for row_num, txn in enumerate(transactions, 2):
            ws.cell(row=row_num, column=1).value = txn.date.isoformat()
            ws.cell(row=row_num, column=2).value = txn.description
            ws.cell(row=row_num, column=3).value = txn.amount
            ws.cell(row=row_num, column=4).value = txn.category
            
            # Format amount column as currency
            ws.cell(row=row_num, column=3).number_format = '#,##0.00'
            
            # Apply borders
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).border = ExcelExporter.BORDER
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="left")
        
        # Total row
        if transactions:
            total_row = len(transactions) + 2
            ws.cell(row=total_row, column=1).value = "TOTAL"
            ws.cell(row=total_row, column=1).font = ExcelExporter.TOTAL_FONT
            ws.cell(row=total_row, column=1).fill = ExcelExporter.TOTAL_FILL
            
            # Sum formula
            total_formula = f"=SUM(C2:C{total_row-1})"
            ws.cell(row=total_row, column=3).value = total_formula
            ws.cell(row=total_row, column=3).font = ExcelExporter.TOTAL_FONT
            ws.cell(row=total_row, column=3).fill = ExcelExporter.TOTAL_FILL
            ws.cell(row=total_row, column=3).number_format = '#,##0.00'
            
            for col in range(1, 5):
                ws.cell(row=total_row, column=col).border = ExcelExporter.BORDER
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_monthly_summary(summary_data: Dict[str, Any]) -> BytesIO:
        """
        Export monthly summary to Excel
        Multi-sheet: Overview + Category Details
        """
        
        wb = openpyxl.Workbook()
        
        # ===== SHEET 1: Monthly Overview =====
        ws_overview = wb.active
        ws_overview.title = "Monthly Summary"
        
        # Column widths
        ws_overview.column_dimensions['A'].width = 12
        ws_overview.column_dimensions['B'].width = 15
        ws_overview.column_dimensions['C'].width = 15
        ws_overview.column_dimensions['D'].width = 15
        
        # Headers
        headers = ["Month", "Total Income", "Total Expenses", "Net Balance"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_overview.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = ExcelExporter.BORDER
        
        # Data rows
        months = summary_data.get("months", [])
        for row_num, month in enumerate(months, 2):
            ws_overview.cell(row=row_num, column=1).value = month["month"]
            ws_overview.cell(row=row_num, column=2).value = month["total_income"]
            ws_overview.cell(row=row_num, column=3).value = month["total_expenses"]
            ws_overview.cell(row=row_num, column=4).value = month["net_balance"]
            
            for col in range(1, 5):
                cell = ws_overview.cell(row=row_num, column=col)
                cell.border = ExcelExporter.BORDER
                if col > 1:
                    cell.number_format = '#,##0.00'
        
        # Overall totals
        if months:
            total_row = len(months) + 2
            overall = summary_data.get("overall", {})
            
            ws_overview.cell(row=total_row, column=1).value = "OVERALL"
            ws_overview.cell(row=total_row, column=2).value = overall.get("total_income", 0)
            ws_overview.cell(row=total_row, column=3).value = overall.get("total_expenses", 0)
            ws_overview.cell(row=total_row, column=4).value = overall.get("net_balance", 0)
            
            for col in range(1, 5):
                cell = ws_overview.cell(row=total_row, column=col)
                cell.fill = ExcelExporter.TOTAL_FILL
                cell.font = ExcelExporter.TOTAL_FONT
                cell.border = ExcelExporter.BORDER
                if col > 1:
                    cell.number_format = '#,##0.00'
        
        # ===== SHEET 2: Category Breakdown =====
        ws_categories = wb.create_sheet("Category Breakdown")
        ws_categories.column_dimensions['A'].width = 20
        ws_categories.column_dimensions['B'].width = 15
        
        # Headers
        for col_num, header in enumerate(["Category", "Total Amount"], 1):
            cell = ws_categories.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = ExcelExporter.BORDER
        
        # Aggregate categories across all months
        all_categories: Dict[str, float] = {}
        for month in months:
            for category, amount in month.get("categories", {}).items():
                if category not in all_categories:
                    all_categories[category] = 0
                all_categories[category] += amount
        
        # Sort by amount descending
        sorted_cats = sorted(all_categories.items(), key=lambda x: x[1], reverse=True)
        
        # Data rows
        for row_num, (category, amount) in enumerate(sorted_cats, 2):
            ws_categories.cell(row=row_num, column=1).value = category
            ws_categories.cell(row=row_num, column=2).value = amount
            ws_categories.cell(row=row_num, column=2).number_format = '#,##0.00'
            
            for col in range(1, 3):
                ws_categories.cell(row=row_num, column=col).border = ExcelExporter.BORDER
        
        # Total row
        if all_categories:
            total_row = len(all_categories) + 2
            ws_categories.cell(row=total_row, column=1).value = "TOTAL"
            ws_categories.cell(row=total_row, column=1).font = ExcelExporter.TOTAL_FONT
            ws_categories.cell(row=total_row, column=1).fill = ExcelExporter.TOTAL_FILL
            
            total_formula = f"=SUM(B2:B{total_row-1})"
            ws_categories.cell(row=total_row, column=2).value = total_formula
            ws_categories.cell(row=total_row, column=2).font = ExcelExporter.TOTAL_FONT
            ws_categories.cell(row=total_row, column=2).fill = ExcelExporter.TOTAL_FILL
            ws_categories.cell(row=total_row, column=2).number_format = '#,##0.00'
            
            for col in range(1, 3):
                ws_categories.cell(row=total_row, column=col).border = ExcelExporter.BORDER
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
