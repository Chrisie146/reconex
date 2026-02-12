"""
Test category export with VAT columns
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from models import SessionLocal
from services.summary import ExcelExporter
import openpyxl

def test_export():
    # Replace with your actual session_id
    session_id = input("Enter your session_id: ").strip()
    
    db = SessionLocal()
    
    try:
        print("\n" + "="*60)
        print("Testing Category Export with VAT")
        print("="*60)
        
        # Test WITHOUT VAT
        print("\n1. Testing export WITHOUT VAT columns...")
        output_no_vat = ExcelExporter.export_all_categories_monthly(
            session_id=session_id,
            db=db,
            include_vat=False
        )
        
        # Load and inspect
        wb = openpyxl.load_workbook(output_no_vat)
        first_sheet = wb.worksheets[0]
        print(f"   Sheet name: {first_sheet.title}")
        
        # Find header row
        for row_idx in range(1, 20):
            cell_value = first_sheet.cell(row_idx, 1).value
            if cell_value == "Date":
                print(f"   Headers at row {row_idx}:")
                headers = []
                for col_idx in range(1, 10):
                    val = first_sheet.cell(row_idx, col_idx).value
                    if val:
                        headers.append(val)
                print(f"   {headers}")
                break
        
        print("   ✓ Export WITHOUT VAT completed")
        
        # Test WITH VAT
        print("\n2. Testing export WITH VAT columns...")
        output_with_vat = ExcelExporter.export_all_categories_monthly(
            session_id=session_id,
            db=db,
            include_vat=True
        )
        
        # Load and inspect
        wb = openpyxl.load_workbook(output_with_vat)
        first_sheet = wb.worksheets[0]
        print(f"   Sheet name: {first_sheet.title}")
        
        # Find header row
        for row_idx in range(1, 20):
            cell_value = first_sheet.cell(row_idx, 1).value
            if cell_value == "Date":
                print(f"   Headers at row {row_idx}:")
                headers = []
                for col_idx in range(1, 10):
                    val = first_sheet.cell(row_idx, col_idx).value
                    if val:
                        headers.append(val)
                print(f"   {headers}")
                break
        
        print("   ✓ Export WITH VAT completed")
        
        print("\n" + "="*60)
        print("RESULTS:")
        print("="*60)
        print("If you see 'Amount (Incl VAT)', 'VAT Amount', 'Amount (Excl VAT)'")
        print("in the second test, then the backend is working correctly!")
        print("If not, there may be an issue with the export logic.")
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

if __name__ == "__main__":
    test_export()
